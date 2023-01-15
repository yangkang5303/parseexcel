from PIL import Image, ImageTk
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import json
import csv

def parse_json_data():
    # Open a file dialog to select an Excel file
    filepath = filedialog.askopenfilename()

    # Open the Excel file
    workbook = load_workbook(filepath)
    worksheet = workbook.active

    # Iterate over the rows and parse the JSON data in column C, ignore the header
    result_string = ""
    results = []
    for row in worksheet.iter_rows(min_row=2, max_col=3, values_only=True):
        try:
            data = json.loads(row[2])
            results.append([row[0], row[1], data['a']])
        except:
            pass
    # Write the results to a new CSV file
    with open("results.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(results)

    # Display the results in a label widget
    result_label.config(text="Data written to results.csv.")

# Create the main window
root = tk.Tk()
root.geometry("1000x500")
root.title("Excel JSON Parser")

# Open the image file
image = Image.open("/Users/yangkang/Desktop/IMG_20211110_084611R_2.jpg")

# Resize the image
image = image.resize((700, 400), Image.ANTIALIAS)

# Create a PhotoImage object
photo_image = ImageTk.PhotoImage(image)

# Create a label to display the image
image_label = tk.Label(root, image=photo_image)
image_label.place(x=10, y=100)

# Create a label to display instructions
instruction_label = tk.Label(root, text="请上传excel:")
instruction_label.pack()

# Create a button to open the file dialog
file_button = tk.Button(root, text="浏览", command=parse_json_data)
file_button.pack()

# Create a label to display the results
result_label = tk.Label(root)
result_label.pack()

# Run the main loop
root.mainloop()
